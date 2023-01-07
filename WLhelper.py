import time
import random
import discum
from openpyxl import load_workbook
from discord.ext import commands
import logging
import requests
import asyncio


class WLHelper:
    def __init__(self):
        self.token1 = 'MzI1MTI4NjIyODUzODQ5MDg5.Ge6uwB.dZCBsF4p5yDnR-V1bpbRi9P6dYasLrr7R1vCKQ'
        self.token2 = ''
        self.id_bot1 = ''
        self.id_bot2 = ''
        self.id_channel = '946492345426260038'
        self.time_min = int
        self.time_max = int

    def solo_bot(self):
        bot = discum.Client(token=self.token1, log={"console": True, "file": False})
        choice_file = open('base_solo.txt')
        line_list = []
        for line in choice_file:
            line_list.append(line)
        choice_file.close()
        while True:
            choice = random.choice(line_list)
            bot.typingAction(channelID=self.id_channel)
            time.sleep(10)
            bot.sendMessage(channelID=self.id_channel, message=choice)
            time.sleep(random.randint(self.time_min, self.time_b))

    def double_bot(self):
        bot1 = discum.Client(token=self.token1, log={"console": True, "file": False})
        bot2 = discum.Client(token=self.token2, log={"console": True, "file": False})
        wb = load_workbook("base_duo.xlsx")
        ws = wb.get_sheet_by_name('Лист1')
        fst_message = ws['A']
        sec_message = ws['B']
        fst_message_list = [fst_message[x].value for x in range(len(fst_message))]
        sec_message_list = [sec_message[x].value for x in range(len(sec_message))]
        bot2.typingAction(channelID=self.id_channel)
        time.sleep(10)
        bot2.sendMessage(channelID=self.id_channel, message=fst_message_list[random.randint(1, len(fst_message_list))])
        while True:
            message_number = random.randint(1, len(fst_message_list))
            time.sleep(7)
            messages_from_bot2 = bot2.searchMessages(channelID=self.id_channel, authorID=self.id_bot2)
            last_mess_bot2 = [i[0]['id'] for i in messages_from_bot2.json()['messages']][0]
            bot1.typingAction(channelID=self.id_channel)
            time.sleep(10)
            bot1.reply(channelID=self.id_channel, messageID=last_mess_bot2, message=sec_message_list[message_number])
            time.sleep(7)
            messages_from_bot1 = bot1.searchMessages(channelID=self.id_channel, authorID=self.id_bot1)
            last_mess_bot1 = [i[0]['id'] for i in messages_from_bot1.json()['messages']][0]
            bot2.typingAction(channelID=self.id_channel)
            time.sleep(10)
            bot2.reply(channelID=self.id_channel, messageID=last_mess_bot1, message=fst_message_list[message_number])
            time.sleep(random.randint(self.time_min, self.time_b))
        bot1.gateway.run(auto_reconnect=True)
        bot2.gateway.run(auto_recconnect=True)

    def ai_bot(self):
        logging.basicConfig(level=logging.INFO)
        ai_bot = commands.Bot(command_prefix='prefix', self_bot=True)
        logging.basicConfig(level=logging.INFO)

        async def get_answer(message):
            url = 'http://api.brainshop.ai/get?bid=171680&key=luYlksAX8v7aTNSZ&uid=1111&msg='
            url += message
            r = requests.get(url)
            return r.json()['cnt']

        @ai_bot.event
        async def on_ready():
            logging.info("The bot is ready.")
            channel = ai_bot.get_channel(int(self.id_channel))
            await channel.send('/старт?')

        @ai_bot.event
        async def on_message(message):
            mention = f'<@!{ai_bot.user.id}>'
            channel = message.channel
            if (message.reference is not None) and (message.author.id != ai_bot.user.id):
                content = await get_answer(message.content)
                async with channel.typing():
                    await asyncio.sleep(2.1)
                    await message.reply(content)
                    print(f"Message '{content}' sent!")

        ai_bot.run(self.token1, bot=False)




